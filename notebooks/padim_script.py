"""
Padim Script
"""

import os
import time
from time import gmtime
from time import strftime

import cv2
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import torch
from sklearn.metrics import roc_auc_score
from torch.utils.data import DataLoader
from torchvision import transforms as T

import anodet


def visual_image_result(test_images, score_map_classifications, image_classifications, score_maps, images, image_scores):
    """
    Visual image result with heatmaps

        Args:
            test_images: Images on which to draw boundaries
            score_map_classifications: Anomaly classifications about the images
            image_classifications: Information about, if the images have anomalies
            score_maps: The values to use to generate colormap
            images: The images to be drawn upon
            image_scores: Image level scores

    """
    boundary_images = anodet.visualization.framed_boundary_images(test_images, score_map_classifications,
                                                                  image_classifications, padding=40)
    heatmap_images = anodet.visualization.heatmap_images(test_images, score_maps, alpha=0.5)
    highlighted_images = anodet.visualization.highlighted_images(images, score_map_classifications, color=(128, 0, 128))

    for i, idx in enumerate(range(len(images))):
        fig, axs = plt.subplots(1, 4, figsize=(12, 6))
        fig.suptitle('Image: ' + str(idx) + " score: " + str(image_scores[i].item()), y=0.75, fontsize=20)
        axs[0].imshow(images[idx])
        axs[1].imshow(boundary_images[idx])
        axs[2].imshow(heatmap_images[idx])
        axs[3].imshow(highlighted_images[idx])
        plt.show()


def write_dict_values_to_sheet(file_path, dict_parameters):
    """
    Write dictionary values to excel sheet

    Args:
        file_path: Path to sheet in .xlsx format
        dict_parameters: Dictionary with k/v pairs.

    """
    # Open sheet
    book = openpyxl.load_workbook(file_path)
    sheet = book.active

    # Write to latest row
    row = sheet.max_row + 1
    for i, val in enumerate(dict_parameters.values()):
        c1 = sheet.cell(row=row, column=i + 2)
        c1.value = str(val)
    book.save(file_path)


def run_padim_script(dataset_path=os.path.realpath('../../data/pscb/good_cropped_masked'),
                     test_good_images_path=os.path.realpath('../../data/pscb/test_good_cropped_masked'),
                     test_bad_images_path=os.path.realpath('../../data/pscb/bad_cropped_masked'),
                     backbone: str = 'resnet34',
                     tresh: int = 11,
                     extractions: int = 1,
                     train_images_limit=None,
                     test_good_images_limit=None,
                     test_bad_images_limit=None,
                     result_sheet_path=None,
                     visualize_result=True,
                     image_transforms: T.Compose = T.Compose([T.Resize(112),
                                                              T.CenterCrop(112),
                                                              T.ToTensor(),
                                                              T.Normalize(mean=[0.485, 0.456, 0.406],
                                                                          std=[0.229, 0.224, 0.225])
                                                              ])
                     ):
    """
    Script to run Padim model using AnodetDataset.

        Args:
            dataset_path: Path for training images.
            test_good_images_path: Path to good test images.
            test_bad_images_path: Path to bad test images.
            backbone: The name of the desired backbone. Must be one of: [resnet18, wide_resnet50]
            tresh: A treshold value. If an image score is larger than
                or equal to thresh it is classified as anomalous.
            extractions: Number of extractions from dataloader. Could be of interest \
                when applying random augmentations.
            train_images_limit: Limit number of images to train on model.
            test_good_images_limit: Limit number of good images to test on model.
            test_bad_images_limit: Limit number of good bad to test on model.
            result_sheet_path: Path to excel sheet with .xlsx extension, leave empty to skip.
            visualize_result: View test images. Visualize results of binary prediction.
            image_transforms: Torchvision Compose obj containing transforms

        Returns:
            image_classification_target: Numpy array with target classification
            image_classifications: Numpy array with predicted image classifcations

    """
    # Start time
    start_time = time.time()

    # Load dataset
    dataset = anodet.AnodetDataset(image_directory_path=dataset_path,
                                   image_transforms=image_transforms,
                                   images_limit=train_images_limit)
    dataloader = DataLoader(dataset, batch_size=8)

    # Init the model
    model = anodet.Padim(backbone=backbone)
    model.fit(dataloader=dataloader, extractions=extractions)

    # Load good and bad test images
    good_image_paths = anodet.get_paths_for_directory_path(
        directory_path=test_good_images_path,
        limit=test_good_images_limit)
    bad_image_paths = anodet.get_paths_for_directory_path(
        directory_path=test_bad_images_path,
        limit=test_bad_images_limit)
    test_paths = good_image_paths + bad_image_paths

    # Target array image level
    image_classification_target = [1] * len(good_image_paths) + [0] * len(bad_image_paths)
    image_classification_target = np.array(image_classification_target)

    images = []
    for path in test_paths:
        image = cv2.imread(path)
        image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
        images.append(image)

    batch = anodet.to_batch(images, image_transforms, torch.device('cpu'))

    # Make prediction
    image_scores, score_maps = model.predict(batch)

    # Interpret the prediction TODO: Classifications on patch level
    score_map_classifications = anodet.classification(score_maps, tresh)
    image_classifications = anodet.classification(image_scores, tresh)
    image_classifications = image_classifications.numpy()

    # Calculate time
    end_time = time.time() - start_time
    end_time = strftime("%H:%M:%S", gmtime(end_time))

    # Visualize result
    if visualize_result:
        test_images = np.array(images).copy()
        visual_image_result(test_images, score_map_classifications, image_classifications, score_maps, images, image_scores)
        anodet.visualize_eval_pair(image_classification_target, image_classifications)

    # Save result to sheet
    if result_sheet_path:
        score_roc_auc = roc_auc_score(image_classification_target,
                                      image_classifications)
        precision, recall, thresholds = anodet.optimal_threshold(image_classification_target,
                                                                 image_classifications)
        sheet_info = {"score_roc_auc": score_roc_auc, "optimal_trest": thresholds, "precision": precision, "recall":
                      recall, "time": end_time, "backbone": backbone, "tresh": tresh, "extractions": extractions,
                      "train_images_limit": train_images_limit, "image_transforms": str(image_transforms)}
        write_dict_values_to_sheet(result_sheet_path, sheet_info)

    return image_classification_target, image_classifications
